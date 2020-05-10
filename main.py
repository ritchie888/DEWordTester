# -*- coding: utf-8 -*-
import os
import sys
import csv
import random
import datetime
from termcolor import colored
from openpyxl import load_workbook
from openpyxl import Workbook


def game(mode, data):
    # Create an empty array to hold the data.
    resultData = []
    for i in range(0, len(data)):
        # The 'question' is always the German word, which is in the 1st column (idx 0).
        question = data[i][0]
        # Get the user's answer, using input.
        usrAns = input("[" + str(i + 1) + " of " + str(len(data)) + "] " + colored(question, "yellow") + ": ")
        # Define the real/correct answer, which is defined by the column specified in 'mode'
        realAns = data[i][mode]
        # If the answer typed is 'exit!' that's the break command to exit before finishing the list of questions.
        if usrAns.lower() == "exit!":
            return resultData
        elif (usrAns.lower() == "") | (usrAns.lower() == " ") | (usrAns.lower() not in realAns.lower()):
            # If the answer is incorrect, print in red. Also print the correct answer in yellow.
            print(colored("Incorrect. ", "red") + colored("Correct answer is \'", "yellow") + colored(realAns, "cyan") + colored("\'.", "yellow"))
            # Append the result in the result array.
            resultData.append([str(i + 1), str(len(data)), question.lower(), realAns.lower(), usrAns.lower(), 0])
        elif usrAns.lower() in realAns.lower():
            # If the answer is correct, print in green.
            print(colored("Correct. ", "green") + colored("Full answer is \'", "yellow") + colored(realAns, "cyan") + colored("\'.", "yellow"))
            # Append the result in the result array.
            resultData.append([str(i + 1), str(len(data)), question.lower(), realAns.lower(), usrAns.lower(), 1])

    # Return the data results once completed/quit.
    return resultData


def readCsv(fileRead):
    # It is important that the encoding parameter is entered here for the German characters.
    with open(fileRead + ".csv", "r", newline="", encoding="utf-8") as csvfile:
        data = list(csv.reader(csvfile))
    # Return the data.
    return data


def readXlsx(fileRead, sheetname):
    # Open file.
    wb = load_workbook(fileRead + '.xlsx')
    # Get the index value of the sheet of interest.
    idx = wb.sheetnames.index(sheetname)
    # Select the sheet using the index.
    sheet = wb.worksheets[idx]
    # Create empty list to hold the data.
    data = []
    # loop through the sheet, and append the rows to the data.
    for row in sheet.values:
        data.append(row)
    # Return the data.
    return data


def writeLocalXlsx(filename, sheetname, data):
    # Create a workbook.
    wb = Workbook()
    # Set the default sheet as the active sheet.
    sheet = wb.active
    # Change its name.
    sheet.title = sheetname
    # Append the data to the sheet.
    for d in data:
        sheet.append(d)
    # Save the sheet and close.
    wb.save(filename=filename + '_results.xlsx')


def writeGlobalXlsx(filename, sheetname, data):
    # If the file doesn't already exist, create.
    if not os.path.isfile(filename):
        # Create a workbook.
        wb = Workbook()
        # Set the default sheet as the active sheet.
        sheet = wb.active
        # Change its name.
        sheet.title = sheetname
    else:
        # If the file already exists, open it.
        wb = load_workbook(filename)
        # Get the list of current worksheets
        sheetlist = wb.sheetnames
        # If the sheet we're trying to write to does not exist, create and append to the end.
        if sheetname not in sheetlist:
            sheet = wb.create_sheet(sheetname, len(sheetname) + 1)
        else:
            # If it does exist, find its index in the list.
            idx = sheetlist.index(sheetname)
            # Point the sheet to the correct index.
            sheet = wb.worksheets[idx]
    # Append the data to the sheet.
    sheet.append(data)
    # Save the sheet and close.
    wb.save(filename=filename)


def calculatePercentage(resultData):
    # Initialise variables to collect the correct answers.
    cor = 0
    # Loop through all answers.
    for i in range(0, len(resultData)):
        cor += resultData[i][-1]
    # Calculate percentage and return.
    return [(cor / len(resultData)) * 100.0, cor, len(resultData)]


# MAIN
if __name__ == "__main__":
    # This is needed to setup coloured text on Windows. Should work automatically on Linux.
    if sys.platform == 'win32':
        os.system('color')

    # Specify the file to read from. Must be csv. If fileRead is blank, then the user can enter the csv name, provided it is located at the same location as this .py file.
    fileRead = "WB_frequencydictionary_jones"
    fileReadShort = "fd_jones"
    sheetRead = "Verbs"
    writeLoc = "DATA/"
    if fileRead == "":
        fileRead = input("Specify name of file to read (without extension): ")
        fileReadShort = input("Specify a shortform name of the file entered: ")
        sheetRead = input("Specify name of sheet within the file to read ")

    # Read the xlsx file.
    data = readXlsx(fileRead, sheetRead)

    # Randomise the list, but keep the row structure.
    random.shuffle(data)

    # Data are organised in two ways: German, English; or German, English, Article
    # Check the size of the first row of data for two or three columns.
    colNum = len(data[1])

    # If else statement to get the user to select which version they want to play.
    while 1:
        if (colNum < 2) | (colNum > 3):
            print(colored("Error: Data loaded is not in the correct format. Check source. Exiting.", "red"))
            exit()
        elif colNum == 2:
            # User does not need to enter a selection as there is only one game mode: English.
            gameMode = 1
            sheetName = fileReadShort + "_" + sheetRead + '_Eng'
            break
        elif colNum == 3:
            # Define game mode. 1 = English words, 2 = articles.
            gameMode = input("Specify game mode (1 = English, 2 = articles): ")
            try:
                if int(gameMode) == 1:
                    sheetName = fileReadShort + "_" + sheetRead + '_Eng'
                    break
                elif int(gameMode) == 2:
                    sheetName = fileReadShort + "_" + sheetRead + '_Art'
                    break
                else:
                    print(colored("Warning: Incorrect selection, try again.", "yellow"))
            except:
                print(colored("Warning: Error occured. Did you enter an integer? Try again.", "yellow"))

    # Grab the time/date, for use in record taking.
    timestamp = '{date:%Y%m%d_%H%M%S}'.format(date=datetime.datetime.now())

    # Initiate game.
    resultData = game(int(gameMode), data)

    # If the user exited before answering any questions, exit here.
    if not resultData:
        exit(0)

    # Create write location if it doesn't exist.
    if not os.path.exists(writeLoc):
        os.makedirs(writeLoc)
        
    # Calculate percentage.
    testScore, corAns, totAns = calculatePercentage(resultData)
    
    # Print result.
    print("\nScore: " + str(testScore) + "% [" + str(corAns) + " / " + str(totAns) + "]")

    # Write local data.
    writeLocalXlsx(writeLoc + timestamp, timestamp, resultData)
    # Calculate percentage and write to global results.
    writeGlobalXlsx(writeLoc + 'GlobalResults.xlsx', sheetName, [timestamp, testScore])
